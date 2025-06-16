# app.py

from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory, flash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, login_required, logout_user, current_user, UserMixin
from flask_dance.contrib.google import make_google_blueprint, google
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from openai import OpenAI
from dotenv import load_dotenv
import os
import base64
import openpyxl
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import datetime

# Load environment variables
load_dotenv()

# Configuration
UPLOAD_FOLDER = 'uploads'
EXCEL_FOLDER = 'excels'
PDF_FOLDER = 'pdfs'
DB_FOLDER = 'db'
os.makedirs(DB_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
GPT_MODEL = "gpt-4-vision-preview"

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev")
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///db/users.db'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['EXCEL_FOLDER'] = EXCEL_FOLDER
app.config['PDF_FOLDER'] = PDF_FOLDER
app.config['SESSION_COOKIE_SECURE'] = True

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXCEL_FOLDER, exist_ok=True)
os.makedirs(PDF_FOLDER, exist_ok=True)

# Initialize DB and Login
db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Google OAuth Setup
google_bp = make_google_blueprint(client_id=os.getenv("GOOGLE_OAUTH_CLIENT_ID"),
                                   client_secret=os.getenv("GOOGLE_OAUTH_CLIENT_SECRET"),
                                   redirect_to="google_login")
app.register_blueprint(google_bp, url_prefix="/login")

# OpenAI Setup
openai = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# User Model
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(150))

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash("Credenziali non valide", "danger")
    return render_template('index.html')

@app.route('/register', methods=['POST'])
def register():
    email = request.form['email']
    password = request.form['password']
    if User.query.filter_by(email=email).first():
        flash("Email già registrata", "warning")
        return redirect(url_for('login'))
    hashed_pw = generate_password_hash(password)
    user = User(email=email, password=hashed_pw)
    db.session.add(user)
    db.session.commit()
    flash("Registrazione completata. Ora puoi accedere.", "success")
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html', user=current_user)

@app.route('/upload_excel', methods=['POST'])
@login_required
def upload_excel():
    file = request.files['excel']
    month_name = request.form['month']
    if file and allowed_file(file.filename):
        safe_filename = secure_filename(file.filename)
        user_folder = os.path.join(EXCEL_FOLDER, current_user.email)
        month_folder = os.path.join(user_folder, month_name)
        os.makedirs(month_folder, exist_ok=True)
        path = os.path.join(month_folder, 'spese_mensili.xlsx')
        file.save(path)
        canvas.Canvas(os.path.join(month_folder, 'scontrini_unificati.pdf'), pagesize=A4).save()
        flash("Nota spese creata!", "success")
    else:
        flash("Formato file non valido", "danger")
    return redirect(url_for('dashboard'))

@app.route('/upload_receipt/<month>', methods=['POST'])
@login_required
def upload_receipt(month):
    img = request.files['receipt']
    if img:
        user_folder = os.path.join(EXCEL_FOLDER, current_user.email, month)
        img_path = os.path.join(user_folder, secure_filename(img.filename))
        img.save(img_path)

        with open(img_path, "rb") as f:
            b64_img = base64.b64encode(f.read()).decode("utf-8")

        response = openai.chat.completions.create(
            model=GPT_MODEL,
            messages=[
                {"role": "system", "content": "Estrai da uno scontrino: data, importo, valuta, negozio, categoria. Rispondi in JSON."},
                {"role": "user", "content": [
                    {"type": "text", "text": "Ecco uno scontrino."},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_img}"}}
                ]},
            ],
            max_tokens=500
        )

        json_data = response.choices[0].message.content

        wb = load_workbook(os.path.join(user_folder, 'spese_mensili.xlsx'))
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                row[0].value = json_data
                riga = row[0].row
                break
        wb.save(os.path.join(user_folder, 'spese_mensili.xlsx'))

        pdf_path = os.path.join(user_folder, 'scontrini_unificati.pdf')
        c = canvas.Canvas(pdf_path, pagesize=A4)
        c.drawImage(img_path, 20, 200, width=400, preserveAspectRatio=True)
        c.drawString(20, 50, f"Riga Excel: {riga}")
        c.showPage()
        c.save()
        flash("Scontrino caricato e analizzato", "success")

    return redirect(url_for('dashboard'))

@app.route('/download/<month>/<filename>')
@login_required
def download_file(month, filename):
    user_folder = os.path.join(EXCEL_FOLDER, current_user.email, month)
    return send_from_directory(user_folder, filename, as_attachment=True)

@app.route('/login/google')
def google_login():
    if not google.authorized:
        return redirect(url_for("google.login"))
    resp = google.get("/oauth2/v2/userinfo")
    if resp.ok:
        email = resp.json()["email"]
        user = User.query.filter_by(email=email).first()
        if not user:
            user = User(email=email)
            db.session.add(user)
            db.session.commit()
        login_user(user)
    return redirect(url_for("dashboard"))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

# Helpers
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# CLI DB init
@app.cli.command("init-db")
def init_db():
    db.create_all()
    print("Database initialized")

if __name__ == '__main__':
    app.run(debug=True)
